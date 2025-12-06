import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
import requests
from bs4 import BeautifulSoup
from datetime import datetime
import json
import re
import io
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill

# Configuration
API_TOKEN = 'ZE1HOU1BcE9UNVk=|1782085439|ygCv5CrNibKxK1yz4beV84m4h6KMVD8Zf2ae/KCP33o='
BASE_URL = 'https://platform.quip-amazon.com'
QUIP_SOURCES = {
    '2025_Q1': {'id': 'tP7kA63aBaaR', 'name': 'JC Q1 2025'},
    '2025_Q2': {'id': 'BbaaAYz9OOQ7', 'name': 'JC Q2 2025'},
    '2025_Q3': {'id': 'b0w8Awc6xyjW', 'name': 'JC Q3 2025'},
    '2025_Q4': {'id': 'o3q3AgLHlYn2', 'name': 'JC Q4 2025'},
    # 2026 Sources
    '2026_Q1': {'id': 'IvhXAG3KcCug', 'name': 'JC Q1 2026'},
    '2026_Q2': {'id': 'SFxdAl8S3zc3', 'name': 'JC Q2 2026'},
    '2026_Q3': {'id': 'Un2jA4qLCYaO', 'name': 'JC Q3 2026'},
    '2026_Q4': {'id': '9RajAh6HjK6a', 'name': 'JC Q4 2026'}
}

def fetch_quip_data(quarter):
    """Fetch data from QUIP"""
    try:
        quip_id = QUIP_SOURCES[quarter]['id']
        headers = {
            'Authorization': f'Bearer {API_TOKEN}',
            'Content-Type': 'application/json'
        }
        
        response = requests.get(
            f'{BASE_URL}/1/threads/{quip_id}',
            headers=headers
        )
        
        if response.status_code == 200:
            return response.json()
        else:
            st.error(f"Failed to fetch QUIP data: {response.status_code}")
            return None
            
    except Exception as e:
        st.error(f"Error connecting to QUIP: {str(e)}")
        return None

def get_available_tables(quip_data):
    """Get list of available tables/tabs from QUIP data"""
    soup = BeautifulSoup(quip_data['html'], 'html.parser')
    tables = soup.find_all('table')
    
    table_titles = []
    for table in tables:
        title = table.get('title', '').strip()
        # Check how tables are titled in 2026 documents
        print(f"Found table title: {title}")  # Add this for debugging
        if title.startswith('WK'):
            table_titles.append(title)
    
    return sorted(table_titles)

def parse_quip_data(quip_data_dict, selected_quarters, selected_weeks):
    """Parse QUIP HTML content into DataFrame for selected quarters and weeks"""
    all_data = []
    timestamp = datetime.now()
    
    for quarter, quip_data in quip_data_dict.items():
        soup = BeautifulSoup(quip_data['html'], 'html.parser')
        tables = soup.find_all('table')
        
        for table in tables:
            table_title = table.get('title', '').strip()
            if table_title in selected_weeks:
                rows = table.find_all('tr')
                # Skip the first row (header row)
                for row in rows[1:]:
                    cells = row.find_all('td')
                    if len(cells) >= 19:
                        row_data = {
                            'Quarter': quarter,
                            'Week': table_title,
                            'Station': cells[1].text.strip(),          
                            'Status': cells[2].text.strip(),           
                            'OFD': cells[3].text.strip(),             
                            'Business_Type': cells[6].text.strip(),                          
                            'Chain': cells[9].text.strip(),          
                            'A Owner': cells[14].text.strip(),        
                            'Category': cells[17].text.strip(),
                            'Type': cells[17].text.strip(),
                            'Timestamp': timestamp
                        }
                        # Additional check to filter out header row values
                        if (row_data['Station'] and 
                            row_data['Status'].lower() != 'status' and 
                            row_data['Type'].lower() != 'type'):
                            all_data.append(row_data)
    
    return pd.DataFrame(all_data)

def create_performance_ranking(df):
    """Create a ranked performance table for A Owners"""
    try:
        # Create pivot table with numeric values only
        pivot_data = pd.pivot_table(
            df,
            values='Station',
            index='A Owner',
            columns='Status',
            aggfunc='count',
            fill_value=0
        ).reset_index()
        
        # Calculate totals (excluding the A Owner column)
        numeric_cols = pivot_data.select_dtypes(include=[np.number]).columns
        pivot_data['Total_Tasks'] = pivot_data[numeric_cols].sum(axis=1)
        
        # Add metrics including Merged and Published status
        pivot_data['Complete_Tasks'] = pivot_data.get('Complete', 0)
        pivot_data['Published'] = pivot_data.get('Published', 0)
        pivot_data['In_Progress'] = pivot_data.get('In Progress', 0)
        pivot_data['Blocked'] = pivot_data.get('Blocked', 0)
        pivot_data['Merged'] = pivot_data.get('Merged', 0)
        
        # Calculate completion rate (including Complete, Merged, and Published as completed)
        pivot_data['Total_Completed'] = (
            pivot_data['Complete_Tasks'] + 
            pivot_data['Merged'] + 
            pivot_data['Published']
        )
        
        pivot_data['Completion_Rate'] = (
            (pivot_data['Total_Completed'] / pivot_data['Total_Tasks'] * 100)
            .round(1)
            .fillna(0)
        )
        
        # Select and rename columns
        ranking_df = pivot_data[[
            'A Owner',
            'Total_Tasks',
            'Total_Completed',
            'Complete_Tasks',
            'Published',
            'Merged',
            'In_Progress',
            'Blocked',
            'Completion_Rate'
        ]].sort_values('Total_Tasks', ascending=False)
        
        # Add rank
        ranking_df.insert(0, 'Rank', range(1, len(ranking_df) + 1))
        
        # Rename columns
        ranking_df.columns = [
            'Rank',
            'Specialist',
            'Total Tasks',
            'Completed+Published+Merged',
            'Completed',
            'Published',
            'Merged',
            'In Progress',
            'Blocked',
            'Completion Rate (%)'
        ]
        
        return ranking_df
        
    except Exception as e:
        st.error(f"Error in performance ranking calculation: {str(e)}")
        return pd.DataFrame()

def display_status_comparison(df, selected_weeks):
    st.subheader("Status Comparison")
    
    # Create overall status distribution chart and get its colors
    status_counts = df['Status'].value_counts().reset_index()
    status_counts.columns = ['Status', 'Count']
    
    fig = px.bar(
        status_counts,
        x='Status',
        y='Count',
        title='Overall Status Distribution',
        color='Status'
    )
    
    fig.update_layout(
        xaxis_title="Status",
        yaxis_title="Count",
        showlegend=True,
        height=500,
        bargap=0.2
    )
    
    # Get the color mapping from the figure
    color_discrete_map = {}
    for trace in fig.data:
        color_discrete_map[trace.name] = trace.marker.color
    
    st.plotly_chart(fig, use_container_width=True)
    
    # Status Summary by Week
    st.subheader("Status Summary by Week")
    summary_table = pd.pivot_table(
        df,
        values='Station',
        index=['Quarter', 'Week'],
        columns='Status',
        aggfunc='count',
        fill_value=0
    ).reset_index()
    
    st.dataframe(summary_table)
    
    # Status Summary by A Owner
    st.subheader("Status Summary by Specialist")
    owner_summary = pd.pivot_table(
        df,
        values='Station',
        index='A Owner',
        columns='Status',
        aggfunc='count',
        fill_value=0
    ).reset_index()
    
    # Calculate total tasks and sort by it
    numeric_cols = owner_summary.select_dtypes(include=[np.number]).columns
    owner_summary['Total Tasks'] = owner_summary[numeric_cols].sum(axis=1)
    owner_summary = owner_summary.sort_values('Total Tasks', ascending=False)
    
    # Add totals row
    totals = owner_summary.sum(numeric_only=True).to_frame('Total').T
    totals.insert(0, 'A Owner', 'Total')
    owner_summary = pd.concat([owner_summary, totals])
    
    # Move 'Total Tasks' column to be the second column
    cols = owner_summary.columns.tolist()
    cols.insert(1, cols.pop(cols.index('Total Tasks')))
    owner_summary = owner_summary[cols]

    # Create styles for each column header
    styles = [
        {'selector': f'th.col{i}', 
         'props': [('background-color', color_discrete_map.get(col, '#ffffff')),
                   ('color', 'black'),
                   ('font-weight', 'bold'),
                   ('text-align', 'center')]}
        for i, col in enumerate(owner_summary.columns)
    ]
    
    # Add basic table styles
    styles.extend([
        {'selector': 'th', 'props': [('text-align', 'center'),
                                     ('padding', '8px')]},
        {'selector': 'td', 'props': [('text-align', 'center'),
                                     ('padding', '8px')]}
    ])
    
    # Display the table with colored headers matching the graph
    st.dataframe(
        owner_summary.style
        .set_table_styles(styles)
        .format({col: '{:,.0f}' for col in numeric_cols}),
        height=400
    )

def display_weekly_details(df, selected_weeks):
    st.subheader("Weekly Details")
    
    for week in selected_weeks:
        week_data = df[df['Week'] == week]
        
        with st.expander(f"Week {week} Details", expanded=True):
            col1, col2 = st.columns(2)
            
            with col1:
                fig_status = px.pie(
                    week_data,
                    names='Status',
                    title=f'Status Distribution - Week {week}'
                )
                st.plotly_chart(fig_status, use_container_width=True)
            
            with col2:
                business_counts = week_data['Type'].value_counts().reset_index()
                business_counts.columns = ['Business_Type', 'Count']
                fig_business = px.bar(
                    business_counts,
                    x='Business_Type',
                    y='Count',
                    title=f'Task Type Distribution - Week {week}'
                )
                fig_business.update_layout(
                    xaxis_title="Task Type",
                    yaxis_title="Count"
                )
                st.plotly_chart(fig_business, use_container_width=True)
            
            # Create pivot table with totals
            pivot_table = week_data.pivot_table(
                index='Type',
                columns='Status',
                aggfunc='size',
                fill_value=0
            )
            
            # Add row totals
            pivot_table['Total'] = pivot_table.sum(axis=1)
            
            # Add column totals
            totals_row = pivot_table.sum().to_frame('Total').T
            pivot_table = pd.concat([pivot_table, totals_row])
            
            # Format the table
            st.dataframe(
                pivot_table.style
                .format("{:,.0f}")
                .set_properties(**{
                    'text-align': 'center'
                })
                .apply(lambda x: ['font-weight: bold' if x.name == 'Total' else '' for _ in x], axis=1)
            )

def display_aowner_details(df):
    st.subheader("Specialist Details")
    
    for owner in df['A Owner'].unique():
        owner_data = df[df['A Owner'] == owner]
        
        with st.expander(f"{owner} Details", expanded=True):
            col1, col2 = st.columns(2)
            
            with col1:
                # Status Distribution Pie Chart
                fig_status = px.pie(
                    owner_data,
                    names='Status',
                    title=f'Status Distribution - {owner}'
                )
                st.plotly_chart(fig_status, use_container_width=True)
            
            with col2:
                # Combined Business Type and Type Distribution Bar Chart
                combined_counts = owner_data.groupby(['Business_Type', 'Type']).size().reset_index(name='Count')
                combined_counts['Combined_Type'] = combined_counts['Business_Type'] + ' - ' + combined_counts['Type']
                fig_combined = px.bar(
                    combined_counts,
                    x='Combined_Type',
                    y='Count',
                    title=f'Business Type and Type Distribution - {owner}',
                    color='Business_Type',
                    labels={'Combined_Type': 'Business Type - Type'}
                )
                fig_combined.update_layout(
                    xaxis_title="Business Type - Type",
                    yaxis_title="Count",
                    xaxis={'categoryorder':'total descending'}
                )
                st.plotly_chart(fig_combined, use_container_width=True)
            
            # Detailed breakdown table
            st.markdown(f"##### Detailed Breakdown for {owner}")
            breakdown = pd.pivot_table(
                owner_data,
                values='Station',
                index=['Business_Type', 'Type'],
                columns='Status',
                aggfunc='count',
                fill_value=0
            )
            
            # Add row totals
            breakdown['Total'] = breakdown.sum(axis=1)
            
            # Add column totals
            totals_row = pd.DataFrame(breakdown.sum()).T
            totals_row.index = pd.MultiIndex.from_tuples([('Total', '')])
            
            # Concatenate the totals row
            breakdown = pd.concat([breakdown, totals_row])
            
            # Reset index and rename columns properly
            breakdown = breakdown.reset_index()
            breakdown.columns.name = None  # Remove column name
            breakdown = breakdown.rename(columns={
                'level_0': 'Business_Type',
                'level_1': 'Type'
            })
            
            # Fill NaN values with empty string
            breakdown = breakdown.fillna('')
            
            # Format the table
            st.dataframe(
                breakdown.style
                .format("{:,.0f}", subset=breakdown.columns[2:])  # Format only numeric columns
                .set_properties(**{
                    'text-align': 'center'
                })
                .apply(lambda x: ['font-weight: bold' if x.name == len(breakdown)-1 else '' for _ in x], axis=1)
            )
            
            # Week by week progress
            st.markdown(f"##### Weekly Progress for {owner}")
            weekly_progress = pd.pivot_table(
                owner_data,
                values='Station',
                index=['Quarter', 'Week'],
                columns='Status',
                aggfunc='count',
                fill_value=0
            ).reset_index()
            
            # Add row totals to weekly progress
            weekly_progress['Total'] = weekly_progress.select_dtypes(include=[np.number]).sum(axis=1)
            
            # Add column totals to weekly progress
            totals_row = pd.DataFrame([['Total', ''] + list(weekly_progress.select_dtypes(include=[np.number]).sum())], 
                                    columns=weekly_progress.columns)
            weekly_progress = pd.concat([weekly_progress, totals_row], ignore_index=True)
            
            st.dataframe(
                weekly_progress.style
                .format("{:,.0f}", subset=weekly_progress.columns[2:])  # Format only numeric columns
                .set_properties(**{
                    'text-align': 'center'
                })
                .apply(lambda x: ['font-weight: bold' if x.name == len(weekly_progress)-1 else '' for _ in x], axis=1)
            )

def display_performance_rankings(df):
    st.subheader("Performance Rankings")
    
    rankings = create_performance_ranking(df)
    
    st.dataframe(
        rankings.style
        .format({
            'Completion Rate (%)': '{:.1f}%',
            'Total Tasks': '{:,.0f}',
            'Completed+Published+Merged': '{:,.0f}',
            'Completed': '{:,.0f}',
            'Published': '{:,.0f}',
            'Merged': '{:,.0f}',
            'In Progress': '{:,.0f}',
            'Blocked': '{:,.0f}'
        })
        .background_gradient(
            subset=['Completion Rate (%)'],
            cmap='RdYlGn',  # Red-Yellow-Green colormap
            vmin=0,
            vmax=100
        )
        .set_properties(**{
            'text-align': 'center',
            'font-size': '1em',
            'padding': '5px'
        })
        .set_table_styles([
            {'selector': 'th',
             'props': [
                 ('font-weight', 'bold'),
                 ('text-align', 'center'),
                 ('padding', '5px'),
                 ('background-color', '#f0f2f6')
             ]},
            {'selector': 'td',
             'props': [('text-align', 'center')]},
        ]),
        height=400
    )
    
    # Add visualization of top performers
    if not rankings.empty:
        st.markdown("### Task Distribution by Specialist")
        fig = px.bar(
            rankings,
            x='Specialist',
            y=['Completed', 'Published', 'Merged', 'In Progress', 'Blocked'],
            title='Task Distribution by Specialist',
            barmode='group',
            color_discrete_sequence=['#2e5cb8', '#00a3bf', '#FF9900', '#4B92DB', '#DC3912']
        )
        
        fig.update_layout(
            xaxis_title="Specialist",
            yaxis_title="Number of Tasks",
            legend_title="Status",
            height=500,
            bargap=0.2,
            bargroupgap=0.1
        )
        
        st.plotly_chart(fig, use_container_width=True)

def display_raw_data(df):
    st.subheader("Raw Data")
    st.dataframe(df)
    
    csv = df.to_csv(index=False)
    st.download_button(
        label="📥 Download Raw Data",
        data=csv,
        file_name=f"raw_data_{datetime.now().strftime('%Y%m%d_%H%M')}.csv",
        mime='text/csv',
    )

def create_comprehensive_report(df, selected_weeks):
    """Create a comprehensive Excel report with tables and visualizations"""
    output = io.BytesIO()
    workbook = Workbook()
    
    # Helper function to add a dataframe to a worksheet
    def add_dataframe_to_worksheet(ws, df, start_row=1, start_col=1):
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start_row):
            for c_idx, value in enumerate(row, start_col):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.value = str(value) if value is not None else ""  # Convert all values to strings
                cell.alignment = Alignment(horizontal='center')
        
        # Format header
        for cell in ws[start_row]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    
    # Status Overview Sheet
    ws = workbook.active
    ws.title = "Status Overview"
    
    # Status Distribution
    status_counts = df['Status'].value_counts().reset_index()
    status_counts.columns = ['Status', 'Count']
    ws.append(['Status Distribution'])
    add_dataframe_to_worksheet(ws, status_counts, start_row=2)
    
    # Status Summary by Week
    ws.append([])  # Add empty row for spacing
    ws.append(['Status Summary by Week'])
    summary_table = pd.pivot_table(
        df,
        values='Station',
        index=['Quarter', 'Week'],
        columns='Status',
        aggfunc='count',
        fill_value=0
    ).reset_index()
    add_dataframe_to_worksheet(ws, summary_table, start_row=ws.max_row + 1)
    
    # Status Summary by A Owner with totals
    ws.append([])
    ws.append(['Status Summary by A Owner'])
    owner_summary = pd.pivot_table(
        df,
        values='Station',
        index='A Owner',
        columns='Status',
        aggfunc='count',
        fill_value=0
    ).reset_index()
    
    # Add totals
    numeric_cols = owner_summary.select_dtypes(include=[np.number]).columns
    owner_summary['Total Tasks'] = owner_summary[numeric_cols].sum(axis=1)
    owner_summary = owner_summary.sort_values('Total Tasks', ascending=False)
    
    # Add totals row
    totals = owner_summary.sum(numeric_only=True).to_frame('Total').T
    totals.insert(0, 'A Owner', 'Total')
    owner_summary = pd.concat([owner_summary, totals])
    
    add_dataframe_to_worksheet(ws, owner_summary, start_row=ws.max_row + 1)
    
    # Weekly Details Sheet
    ws = workbook.create_sheet("Weekly Details")
    for week in selected_weeks:
        week_data = df[df['Week'] == week]
        ws.append([f"Week {week} Details"])
        
        # Status breakdown
        ws.append(['Status Distribution'])
        status_pivot = week_data.pivot_table(
            index='Type',
            columns='Status',
            aggfunc='size',
            fill_value=0
        ).reset_index()
        
        # Add totals
        status_pivot['Total'] = status_pivot.select_dtypes(include=[np.number]).sum(axis=1)
        totals_row = pd.DataFrame([['Total'] + list(status_pivot.select_dtypes(include=[np.number]).sum())],
                                columns=status_pivot.columns)
        status_pivot = pd.concat([status_pivot, totals_row], ignore_index=True)
        
        add_dataframe_to_worksheet(ws, status_pivot, start_row=ws.max_row + 1)
        ws.append([])  # Add empty row between weeks
    
    # A Owner Details Sheet
    ws = workbook.create_sheet("A Owner Details")
    for owner in df['A Owner'].unique():
        owner_data = df[df['A Owner'] == owner]
        ws.append([f"A Owner: {owner}"])
        
        # Status Distribution
        ws.append(['Status Distribution'])
        status_pivot = owner_data.pivot_table(
            index=['Business_Type', 'Type'],
            columns='Status',
            aggfunc='count',
            values='Station',
            fill_value=0
        ).reset_index()
        
        # Add totals
        status_pivot['Total'] = status_pivot.select_dtypes(include=[np.number]).sum(axis=1)
        totals_row = pd.DataFrame([['Total', ''] + list(status_pivot.select_dtypes(include=[np.number]).sum())],
                                columns=status_pivot.columns)
        status_pivot = pd.concat([status_pivot, totals_row], ignore_index=True)
        
        add_dataframe_to_worksheet(ws, status_pivot, start_row=ws.max_row + 1)
        
        # Weekly Progress
        ws.append([])
        ws.append(['Weekly Progress'])
        weekly_progress = pd.pivot_table(
            owner_data,
            values='Station',
            index=['Quarter', 'Week'],
            columns='Status',
            aggfunc='count',
            fill_value=0
        ).reset_index()
        add_dataframe_to_worksheet(ws, weekly_progress, start_row=ws.max_row + 1)
        
        ws.append([])  # Add empty row between owners
    
    # Performance Rankings Sheet
    ws = workbook.create_sheet("Performance Rankings")
    rankings = create_performance_ranking(df)
    add_dataframe_to_worksheet(ws, rankings)
    
    # Format all sheets
    for sheet in workbook.sheetnames:
        ws = workbook[sheet]
        # Adjust column widths
        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width
    
    workbook.save(output)
    output.seek(0)
    
    return output

def save_snapshot(df, description=""):
    """Save current data as a snapshot"""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    
    # Create snapshot data
    snapshot = {
        'timestamp': timestamp,
        'description': description,
        'data': df.to_dict(),
        'metrics': {
            'total_tasks': len(df),
            'status_counts': df['Status'].value_counts().to_dict(),
            'by_owner': df.groupby('A Owner')['Status'].value_counts().to_dict()
        }
    }
    
    # Initialize snapshots in session state if not exists
    if 'snapshots' not in st.session_state:
        st.session_state.snapshots = {}
    
    # Save snapshot
    st.session_state.snapshots[timestamp] = snapshot
    return timestamp

def compare_snapshots(snapshot1_id, snapshot2_id):
    """Compare two snapshots and return differences"""
    if 'snapshots' not in st.session_state:
        return None
    
    s1 = st.session_state.snapshots.get(snapshot1_id)
    s2 = st.session_state.snapshots.get(snapshot2_id)
    
    if not s1 or not s2:
        return None
    
    # Convert snapshot data back to DataFrames
    df1 = pd.DataFrame.from_dict(s1['data'])
    df2 = pd.DataFrame.from_dict(s2['data'])
    
    # Calculate differences
    comparison = {
        'timestamp1': s1['timestamp'],
        'timestamp2': s2['timestamp'],
        'description1': s1['description'],
        'description2': s2['description'],
        
        # Overall metrics
        'total_tasks_diff': len(df2) - len(df1),
        
        # Status changes
        'status_changes': {
            status: {
                'before': len(df1[df1['Status'] == status]),
                'after': len(df2[df2['Status'] == status]),
                'diff': len(df2[df2['Status'] == status]) - len(df1[df1['Status'] == status])
            }
            for status in set(df1['Status'].unique()) | set(df2['Status'].unique())
        },
        
        # Changes by owner and status
        'owner_status_changes': {
            owner: {
                status: {
                    'before': len(df1[(df1['A Owner'] == owner) & (df1['Status'] == status)]),
                    'after': len(df2[(df2['A Owner'] == owner) & (df2['Status'] == status)]),
                    'diff': len(df2[(df2['A Owner'] == owner) & (df2['Status'] == status)]) - 
                           len(df1[(df1['A Owner'] == owner) & (df1['Status'] == status)])
                }
                for status in set(df1['Status'].unique()) | set(df2['Status'].unique())
            }
            for owner in set(df1['A Owner'].unique()) | set(df2['A Owner'].unique())
        }
    }
    
    return comparison

def format_change(x, is_numeric=True):
    if not is_numeric:
        return x
    try:
        if pd.isna(x) or x == 0:
            return "-"
        return f"{'+' if float(x) > 0 else ''}{x:,.0f}"
    except (ValueError, TypeError):
        return x


def display_comparison_results(comparison):
    """Display snapshot comparison results"""
    if not comparison:
        st.error("Unable to compare snapshots")
        return

    def custom_text_color(value):
        """Return custom text color based on value"""
        if pd.isna(value):
            return ''
        elif value > 0:
            return 'color: #006400'  # Dark green
        elif value < 0:
            return 'color: #8B0000'  # Dark red
        return ''
    
    st.subheader("Snapshot Comparison")
    
    # Display snapshot information
    col1, col2 = st.columns(2)
    with col1:
        st.markdown(f"**Snapshot 1:** {comparison['timestamp1']}")
        if comparison['description1']:
            st.markdown(f"*Description: {comparison['description1']}*")
    with col2:
        st.markdown(f"**Snapshot 2:** {comparison['timestamp2']}")
        if comparison['description2']:
            st.markdown(f"*Description: {comparison['description2']}*")
    
    # Calculate completed tasks (Complete + Merged + Published)
    completed_before = sum(
        comparison['status_changes'][status]['before']
        for status in ['Complete', 'Merged', 'Published']
        if status in comparison['status_changes']
    )
    completed_after = sum(
        comparison['status_changes'][status]['after']
        for status in ['Complete', 'Merged', 'Published']
        if status in comparison['status_changes']
    )
    completed_change = completed_after - completed_before
    
    # Overall changes
    st.markdown("### Overall Changes")
    st.metric("Tasks Completed (Complete + Merged + Published)", completed_after, delta=completed_change)
    
    # Status changes
    st.markdown("### Status Changes")
    status_changes = pd.DataFrame([
        {
            'Status': status,
            'Before': data['before'],
            'After': data['after'],
            'Change': data['diff']
        }
        for status, data in comparison['status_changes'].items()
    ])
    
    st.dataframe(
        status_changes.style
        .format({
            'Before': '{:,.0f}',
            'After': '{:,.0f}',
            'Change': '{:+,.0f}'
        })
        .apply(lambda x: [''] * len(x) if x.name != 'Change' 
               else [custom_text_color(v) for v in x])
        .set_properties(**{
            'text-align': 'center'
        })
    )
    
    # Create pivot table for current values (After)
    pivot_current = pd.pivot_table(
        pd.DataFrame([
            {
                'Owner': owner,
                'Status': status,
                'Count': data['after']
            }
            for owner, status_data in comparison['owner_status_changes'].items()
            for status, data in status_data.items()
        ]),
        values='Count',
        index='Owner',
        columns='Status',
        fill_value=0
    ).reset_index()
    
    # Calculate total for current values
    numeric_cols = pivot_current.select_dtypes(include=[np.number]).columns
    pivot_current['Total'] = pivot_current[numeric_cols].sum(axis=1)
    
    # Changes by Owner
    st.markdown("### Changes by Owner")
    
    # Get unique owners - corrected version
    owners = sorted(comparison['owner_status_changes'].keys())
    
    for owner in owners:
        with st.expander(f"📊 {owner}", expanded=False):
            # Create owner specific data
            owner_data = comparison['owner_status_changes'].get(owner, {})
            
            # Convert owner data to DataFrame
            owner_df = pd.DataFrame([
                {
                    'Status': status,
                    'Before': data['before'],
                    'After': data['after'],
                    'Change': data['diff']
                }
                for status, data in owner_data.items()
                if data['before'] > 0 or data['after'] > 0  # Only show statuses with data
            ])
            
            if not owner_df.empty:
                # Calculate totals
                totals = {
                    'Before': owner_df['Before'].sum(),
                    'After': owner_df['After'].sum(),
                    'Change': owner_df['Change'].sum()
                }
                
                # Display metrics
                col1, col2, col3 = st.columns(3)
                with col1:
                    st.metric("Total Tasks", totals['After'], delta=totals['Change'])
                with col2:
                    completed = owner_df[owner_df['Status'].isin(['Complete', 'Merged', 'Published'])]
                    completed_total = completed['After'].sum() if not completed.empty else 0
                    completed_change = completed['Change'].sum() if not completed.empty else 0
                    st.metric("Complete + Merged + Published", completed_total, delta=completed_change)
                with col3:
                    in_progress = owner_df[owner_df['Status'].isin(['In progress', 'Blocked'])]
                    in_progress_total = in_progress['After'].sum() if not in_progress.empty else 0
                    in_progress_change = in_progress['Change'].sum() if not in_progress.empty else 0
                    st.metric("In Progress/Blocked", in_progress_total, delta=in_progress_change)
                
                 # Display detailed table
                st.markdown("#### Status Details")
                st.dataframe(
                    owner_df.style
                    .format({
                        'Before': '{:,.0f}',
                        'After': '{:,.0f}',
                        'Change': '{:+,.0f}'
                    })
                    .apply(lambda x: [''] * len(x) if x.name != 'Change' 
                           else [custom_text_color(v) for v in x])
                    .set_properties(**{
                        'text-align': 'center',
                        'font-size': '1em',
                        'padding': '5px'
                    })
                    .set_table_styles([
                        {'selector': 'th',
                         'props': [
                             ('font-weight', 'bold'),
                             ('text-align', 'center'),
                             ('padding', '5px'),
                             ('background-color', '#f0f2f6')
                         ]},
                        {'selector': 'td',
                         'props': [('text-align', 'center')]},
                    ])
                )
                
                # Create a bar chart showing changes
                fig = go.Figure()
                fig.add_trace(go.Bar(
                    name='Before',
                    x=owner_df['Status'],
                    y=owner_df['Before'],
                    marker_color='#1f77b4'  # Darker blue
                ))
                fig.add_trace(go.Bar(
                    name='After',
                    x=owner_df['Status'],
                    y=owner_df['After'],
                    marker_color='#7cc7ff'  # Lighter blue
                ))
                
                fig.update_layout(
                    title=f"Status Distribution for {owner}",
                    barmode='group',
                    height=400,
                    bargap=0.15,
                    bargroupgap=0.1,
                    xaxis_title="Status",
                    yaxis_title="Number of Tasks"
                )
                
                st.plotly_chart(fig, use_container_width=True)
            else:
                st.info("No changes to display")
    
    # Summary tables - without columns
    st.markdown("### Summary Tables")
    
    # Current Numbers table at full width
    st.markdown("**Current Numbers**")
    st.dataframe(
        pivot_current.style
        .format({
            col: '{:,.0f}' if col != 'Owner' else str
            for col in pivot_current.columns
        })
        .set_properties(**{
            'text-align': 'center',
            'font-size': '1em',
            'padding': '5px'
        })
        .set_table_styles([
            {'selector': 'th',
             'props': [
                 ('font-weight', 'bold'),
                 ('text-align', 'center'),
                 ('padding', '5px'),
                 ('background-color', '#f0f2f6')
             ]},
            {'selector': 'td',
             'props': [('text-align', 'center')]},
        ]),
        height=400,
        use_container_width=True  # This makes the table use full width
    )

def main():
    st.set_page_config(page_title="JC Tracker Data Analysis Tool", layout="wide", page_icon="📊")
    
    # Updated CSS with better z-index and positioning
    st.markdown("""
        <style>
        .main-title {
            color: #232F3E;
            font-size: 2.5em;
            font-weight: bold;
            text-align: center;
            padding: 20px 0;
            position: relative;
            z-index: 1000;
            background-color: transparent;
        }
        .subtitle {
            color: #666666;
            text-align: center;
            font-style: italic;
            margin-bottom: 30px;
            position: relative;
            z-index: 1000;
            background-color: transparent;
        }
        div[data-testid="stDecoration"] {
            display: none;
        }
        </style>
    """, unsafe_allow_html=True)
    
    st.markdown('<div class="main-title">JC Tracker Data Analysis Tool</div>', unsafe_allow_html=True)
    st.markdown('<div class="subtitle">Track, analyze, and compare status across weeks</div>', unsafe_allow_html=True)
    
    # Data controls in sidebar
    st.sidebar.header("Data Controls")
    
    # Quarter selection with Select All option
    all_quarters = list(QUIP_SOURCES.keys())
    
    # Add "Select All Quarters" checkbox
    if 'select_all_quarters' not in st.session_state:
        st.session_state.select_all_quarters = False
    
    select_all_quarters = st.sidebar.checkbox("Select All Quarters", 
                                            value=st.session_state.select_all_quarters,
                                            key='select_all_quarters_checkbox')
    
    if select_all_quarters:
        selected_quarters = all_quarters
    else:
        selected_quarters = st.sidebar.multiselect(
            'Select Quarter(s)',
            all_quarters,
            key='quarter_selector'
        )
    
    if selected_quarters:
        load_button = st.sidebar.button('Load Data')
        
        if load_button:
            with st.spinner('Fetching data from QUIP...'):
                all_tables = []
                quip_data_dict = {}
                
                for quarter in selected_quarters:
                    quip_data = fetch_quip_data(quarter)
                    if quip_data:
                        quip_data_dict[quarter] = quip_data
                        all_tables.extend(get_available_tables(quip_data))
                
                if quip_data_dict:
                    st.session_state.quip_data_dict = quip_data_dict
                    st.session_state.available_tables = sorted(list(set(all_tables)))
                    st.session_state.data_loaded = True
                    st.sidebar.success('✅ QUIP data fetched successfully!')
                else:
                    st.error('❌ Failed to fetch QUIP data')
                    st.session_state.data_loaded = False
                    return
        
       # Week selector with Select All option (only show if data is loaded)
        if st.session_state.get('data_loaded', False):
            available_weeks = st.session_state.available_tables
            
            # Add "Select All Weeks" checkbox
            if 'select_all_weeks' not in st.session_state:
                st.session_state.select_all_weeks = False
            
            select_all_weeks = st.sidebar.checkbox("Select All Weeks", 
                                                 value=st.session_state.select_all_weeks,
                                                 key='select_all_weeks_checkbox')
            
            if select_all_weeks:
                selected_weeks = available_weeks
            else:
                selected_weeks = st.sidebar.multiselect(
                    'Select Weeks to Compare',
                    available_weeks,
                    default=[available_weeks[-1]]
                )
            
            if selected_weeks:
                df = parse_quip_data(st.session_state.quip_data_dict, selected_quarters, selected_weeks)
                
                # Add A Owner filter in sidebar
                st.sidebar.markdown("---")
                all_owners = sorted(df['A Owner'].unique())
                selected_owners = st.sidebar.multiselect(
                    'Filter by A Owner(s)',
                    options=all_owners,
                    default=all_owners,
                    key='global_owner_filter'
                )
                
                # Filter the dataframe based on selected owners
                filtered_df = df[df['A Owner'].isin(selected_owners)]
                
                # Add snapshot management to sidebar
                st.sidebar.markdown("---")
                st.sidebar.subheader("Snapshot Management")
                
                # Save snapshot
                snapshot_description = st.sidebar.text_input("Snapshot Description (optional)")
                if st.sidebar.button("Save Current Snapshot"):
                    snapshot_id = save_snapshot(filtered_df, snapshot_description)
                    st.sidebar.success(f"Snapshot saved! ID: {snapshot_id}")
                
                # Compare snapshots option
                show_comparison = False
                if 'snapshots' in st.session_state and len(st.session_state.snapshots) >= 2:
                    st.sidebar.markdown("### Compare Snapshots")
                    snapshot_ids = list(st.session_state.snapshots.keys())
                    
                    col1, col2 = st.sidebar.columns(2)
                    with col1:
                        snapshot1 = st.selectbox("First Snapshot", snapshot_ids, index=0)
                    with col2:
                        snapshot2 = st.selectbox("Second Snapshot", snapshot_ids, index=len(snapshot_ids)-1)
                    
                    if st.sidebar.button("Compare Snapshots"):
                        show_comparison = True
                
                # Create tabs
                if show_comparison:
                    tab1, tab2, tab3, tab4, tab5, tab6 = st.tabs([
                        "Status Comparison",
                        "Weekly Details",
                        "Specialist Details",
                        "Performance Rankings",
                        "Raw Data",
                        "Snapshot Comparison"
                    ])
                else:
                    tab1, tab2, tab3, tab4, tab5 = st.tabs([
                        "Status Comparison",
                        "Weekly Details",
                        "Specialist Details",
                        "Performance Rankings",
                        "Raw Data"
                    ])
                
                # Display content in tabs
                with tab1:
                    display_status_comparison(filtered_df, selected_weeks)
                
                with tab2:
                    display_weekly_details(filtered_df, selected_weeks)
                
                with tab3:
                    display_aowner_details(filtered_df)
                
                with tab4:
                    display_performance_rankings(filtered_df)
                
                with tab5:
                    display_raw_data(filtered_df)
                
                # Show comparison results if requested
                if show_comparison:
                    with tab6:
                        comparison = compare_snapshots(snapshot1, snapshot2)
                        if comparison:
                            display_comparison_results(comparison)
                
                # Add comprehensive report download option
                st.sidebar.markdown("---")
                st.sidebar.subheader("Download Comprehensive Report")
                if st.sidebar.button("Generate Comprehensive Report"):
                    with st.spinner('Preparing comprehensive report...'):
                        try:
                            report = create_comprehensive_report(filtered_df, selected_weeks)
                            st.sidebar.download_button(
                                label="📥 Download Comprehensive Report",
                                data=report,
                                file_name=f"comprehensive_report_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )
                            st.sidebar.success('✅ Report ready for download!')
                        except Exception as e:
                            st.sidebar.error(f'Error generating report: {str(e)}')
            else:
                st.info('Please select at least one week to analyze.')
    else:
        st.info('👆 Please select quarter(s) and click "Load Quarter Data" to begin analysis.')
    
    # Footer
    st.markdown("""---""")
    st.markdown(
        """
        <div style='text-align: center'>
            <p>JC Tracker Data Analysis Tool v1.0 | Last Updated: October 2025</p>
        </div>
        """,
        unsafe_allow_html=True
    )

if __name__ == "__main__":
    main()

    
